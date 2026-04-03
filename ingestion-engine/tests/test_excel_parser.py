"""Tests for Excel/CSV parser and template generator."""

import os
import tempfile

import openpyxl
import pytest

from app.importers.excel_parser import parse_excel, rows_to_raw_assets
from app.importers.templates import ASSET_HEADERS, generate_asset_template


def _create_test_excel(rows: list[list]) -> str:
    """Create a temporary .xlsx file with the given rows.

    First row is treated as headers. Returns the file path.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    wb.close()
    return path


class TestParseValidExcel:
    def test_parse_valid_excel(self):
        """Two valid rows should parse with total=2, valid=2, error=0."""
        headers = ["Asset Tag", "Name", "Type", "Vendor", "Model"]
        row1 = ["SVR-001", "web-server-01", "server", "Dell", "R740"]
        row2 = ["SVR-002", "db-server-01", "storage", "HP", "DL380"]
        path = _create_test_excel([headers, row1, row2])

        try:
            result = parse_excel(path)
            assert result.total_rows == 2
            assert len(result.valid_rows) == 2
            assert len(result.error_rows) == 0
            assert result.valid_rows[0].data["asset_tag"] == "SVR-001"
            assert result.valid_rows[1].data["name"] == "db-server-01"
        finally:
            os.unlink(path)


class TestParseExcelWithErrors:
    def test_parse_excel_with_errors(self):
        """1 valid + 2 invalid rows: errors should be detected."""
        headers = ["Asset Tag", "Name", "Type"]
        valid_row = ["SVR-001", "web-server-01", "server"]
        missing_name = ["SVR-002", None, "server"]
        missing_tag_and_type = [None, "orphan-server", None]
        path = _create_test_excel([headers, valid_row, missing_name, missing_tag_and_type])

        try:
            result = parse_excel(path)
            assert result.total_rows == 3
            assert len(result.valid_rows) == 1
            assert len(result.error_rows) == 2
            # First error row is missing name
            assert any(
                "name" in err for err in result.error_rows[0].errors
            )
            # Second error row is missing asset_tag and type
            assert len(result.error_rows[1].errors) >= 2
        finally:
            os.unlink(path)


class TestRowsToRawAssets:
    def test_rows_to_raw_assets(self):
        """Convert parsed rows and verify fields/attributes split."""
        headers = ["Asset Tag", "Name", "Type", "Vendor", "custom_field"]
        row1 = ["SVR-001", "web-server-01", "server", "Dell", "custom_value"]
        path = _create_test_excel([headers, row1])

        try:
            result = parse_excel(path)
            raw_assets = rows_to_raw_assets(result.valid_rows)

            assert len(raw_assets) == 1
            asset = raw_assets[0]
            assert asset.source == "excel"
            assert asset.fields["asset_tag"] == "SVR-001"
            assert asset.fields["name"] == "web-server-01"
            assert asset.fields["type"] == "server"
            assert asset.fields["vendor"] == "Dell"
            # custom_field should go to attributes
            assert asset.attributes is not None
            assert asset.attributes["custom_field"] == "custom_value"
            # unique_key should be asset_tag (no serial_number)
            assert asset.unique_key == "SVR-001"
        finally:
            os.unlink(path)


class TestGenerateTemplate:
    def test_generate_template(self):
        """Verify header values and Notes sheet exists."""
        buf = generate_asset_template()
        wb = openpyxl.load_workbook(buf)

        # Check Assets sheet headers
        ws = wb["Assets"]
        header_values = [ws.cell(row=1, column=i).value for i in range(1, len(ASSET_HEADERS) + 1)]
        expected_display_names = [h[1] for h in ASSET_HEADERS]
        assert header_values == expected_display_names

        # Check example row exists
        example_values = [ws.cell(row=2, column=i).value for i in range(1, len(ASSET_HEADERS) + 1)]
        expected_examples = [h[2] for h in ASSET_HEADERS]
        assert example_values == expected_examples

        # Check Notes sheet exists
        assert "Notes" in wb.sheetnames
        notes = wb["Notes"]
        assert notes.cell(row=1, column=1).value == "Field Name"

        wb.close()
