"""Asset import template generator."""

import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ASSET_HEADERS: list[tuple[str, str, str]] = [
    ("asset_tag", "Asset Tag", "SVR-001"),
    ("name", "Name", "web-server-01"),
    ("type", "Type", "server"),
    ("sub_type", "Sub Type", "rack-mount"),
    ("status", "Status", "inventoried"),
    ("bia_level", "BIA Level", "normal"),
    ("vendor", "Vendor", "Dell"),
    ("model", "Model", "PowerEdge R740"),
    ("serial_number", "Serial Number", "SN123456789"),
    ("property_number", "Property Number", "PROP-2024-001"),
    ("control_number", "Control Number", "CTRL-2024-001"),
]

VALID_VALUES = {
    "type": "server, network, storage, power",
    "status": "inventoried, deployed, operational, maintenance, decommissioned",
    "bia_level": "critical, important, normal, minor",
}


def generate_asset_template() -> io.BytesIO:
    """Create an Excel template workbook for asset import.

    Returns a BytesIO buffer containing the .xlsx file with:
    - A data sheet with styled headers and an example row
    - A Notes sheet with field descriptions and valid values
    """
    wb = openpyxl.Workbook()

    # --- Data sheet ---
    ws = wb.active
    ws.title = "Assets"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, (field_name, display_name, example) in enumerate(ASSET_HEADERS, start=1):
        # Header row
        cell = ws.cell(row=1, column=col_idx, value=display_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

        # Example data row
        ws.cell(row=2, column=col_idx, value=example)

        # Auto-width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(
            len(display_name) + 4, len(example) + 4
        )

    # --- Notes sheet ---
    notes = wb.create_sheet("Notes")
    notes_header_font = Font(bold=True, color="FFFFFF")
    notes_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    notes_headers = ["Field Name", "Display Name", "Required", "Description", "Valid Values"]
    for col_idx, header in enumerate(notes_headers, start=1):
        cell = notes.cell(row=1, column=col_idx, value=header)
        cell.font = notes_header_font
        cell.fill = notes_header_fill

    required_fields = {"asset_tag", "name", "type"}

    for row_idx, (field_name, display_name, example) in enumerate(ASSET_HEADERS, start=2):
        notes.cell(row=row_idx, column=1, value=field_name)
        notes.cell(row=row_idx, column=2, value=display_name)
        notes.cell(row=row_idx, column=3, value="Yes" if field_name in required_fields else "No")
        notes.cell(row=row_idx, column=4, value=f"Example: {example}")
        notes.cell(row=row_idx, column=5, value=VALID_VALUES.get(field_name, ""))

    # Set column widths for Notes
    notes.column_dimensions["A"].width = 20
    notes.column_dimensions["B"].width = 20
    notes.column_dimensions["C"].width = 10
    notes.column_dimensions["D"].width = 30
    notes.column_dimensions["E"].width = 50

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
